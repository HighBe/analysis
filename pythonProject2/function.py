import os
import xlrd
import string
import json
from openpyxl import load_workbook,Workbook
import openpyxl as op

workingPath = ""


# 批量生成用来修改的html文件对应的txt文件，之后用下一个函数将.html后缀改为.txt。
#用来生成之前的记录在表格中的内容
def generateHtmlold():

    f1 = open("/home/liu/桌面/gumtree_tmp/special_html/1.txt", 'r', encoding='utf-8')
    content = f1.read()
    for i in range(50, 101):
        with open("/home/liu/桌面/gumtree_tmp/special_html/" + str(i) + ".txt", "w", encoding='utf-8') as f:
            f.write(content)

    stra = "<html>\n<body>\n<a href=\"https://github.com/"
    strb = "\" style=\"margin-left:50px;\">所在commit地址</a>\n<code>\n<pre style=\"font-size: 20px;font-family:'Times New Roman', Times, serif;color:brown;\">\n"
    strc = "\n</pre>\n</code>\n<p>message:</p>\n<p>李蓝天：改动分类：原因：</p>\n<p>刘志浩：改动分类：原因：</p>\n</body>\n</html>"

    data = xlrd.open_workbook("/home/liu/桌面/gumtree_tmp/newanalysis.xlsx")  # 读取存放分析结果的xlsx文件
    fspecial = open("/home/liu/PycharmProjects/pythonProject2/special1.txt", "r", encoding='utf-8')  # 打开存放commit连接的文件
    s1 = fspecial.readlines()  # 按行读取存放commit连接的文件，每一行有一个连接
    table = data.sheets()[6]  # 读取第六张表的数据
    rows = table.nrows  # 获取行数
    cols = table.col_values(0)  # 读取第一列的内容
    ###从第一行读到最后一行，读取第二列的数据
    for i in range(0, rows):
        cols[i] = int(cols[i])  # 数组中的内容原本是浮点数类型，转换为整型
        fpatch = open("/home/liu/PycharmProjects/GenerateAST/special/" + str(cols[i]) + ".txt", 'r',
                      encoding='utf-8')  # 读取对应的patch文件
        pacth_content = fpatch.read()  # 获得patch中的内容
        # 删除最后多余的sha，最后形式是 用户名/仓库名/commit/sha
        link1, link2, link3 = s1[cols[i]].rpartition('/')
        s1[cols[i]] = link1
        # 将各部分连接起来
        strall = stra + s1[cols[i]] + strb + pacth_content + strc
        fsave = open("/home/liu/桌面/gumtree_tmp/special_html_test/" + str(i) + ".txt", "w", encoding="utf-8")  # 存文件
        fsave.write(strall)



###### 批量修改文件名称（将.txt变为.html）
def txttohtml():
    path = "/home/liu/桌面/gumtree_tmp/special_html_test"
    os.chdir(path)
    files = os.listdir(path)

    for filename in files:
        portion = os.path.splitext(filename)  # 分离文件名与扩展名
        # 如果后缀是.txt
        if portion[1] == '.txt':
            # 重新组合文件名和后缀名
            newname = portion[0] + '.html'  # 修改为.html
            # newname = "ss.html"
            os.rename(filename, newname)

# 批量生成用来修改的html文件对应的txt文件，之后用下一个函数将.html后缀改为.txt。
def generateHtml(pathstr,fLinkstr):
    path = pathstr
    os.chdir(path)
    files = os.listdir(path)
    fLink = open(fLinkstr,"r",encoding="utf-8")
    s1 = fLink.readlines()
    stra = "<html>\n<body>\n<a href=\""
    strb = "\" style=\"margin-left:50px;\">所在commit地址</a>\n<code>\n<pre style=\"font-size: 20px;font-family:'Times New Roman', Times, serif;color:brown;\">\n"
    strc = "\n</pre>\n</code>\n<p>message:</p>\n<p>李蓝天：改动分类：原因：</p>\n<p>梁叶剑：改动分类：原因：</p>\n<p>刘志浩：改动分类：原因：</p>\n</body>\n</html>"
    os.mkdir("webpage")
    for filename in files:
        if (filename == "webpage") or (filename == "saveLink.txt"):
            continue
        strall = ""
        portion = os.path.splitext(filename)#分离文件名与扩展名
        ftmp = open(filename,"r",encoding='utf-8')
        tmpnum = int(portion[0])
        strall = stra + s1[tmpnum] + strb + ftmp.read() + strc
        savefile = open(pathstr + "/webpage/" + str(tmpnum) + ".txt",'w',encoding='utf-8')
        savefile.write(strall)



# 批量生成网址
def generateLink():
    for i in range(0, 222):
        str1 = "https://highbe.github.io/SolidityWorm/HtmlPlace/" + str(i) + ".html\n"
        print(str1)

#将当前文件夹下的文件对应的网址找到并存到一个文件中
def collectLink(pathstr,fLinkstr):
    fLink = open(fLinkstr,'r',encoding='utf-8')#存放所有地址的文件
    s1 = fLink.readlines()
    path = pathstr#分类后的文件夹
    os.chdir(path)
    files = os.listdir(path)
    files.sort(key=lambda x:int(x.split('.')[0]))#读取文件后按照文件名称排序,要求文件名要格式一致

    content = ""#存放所有需要的网址
    for filename in files:
        portion = os.path.splitext(filename)#分离文件名与扩展名,portion[0]是文件名，portion[1]是扩展名
        tmpnum = int(portion[0])
        content += portion[0] + " " + s1[tmpnum] + "\n"

    fsave = open("saveLink.txt","w",encoding="utf-8")
    fsave.write(content)

# class Stack:
#     def __init__(self):
#         self.stack = []
#
#     def push(self,val):
#         self.stack.append(val)
#
#     def pop(self):
#         return self.stack.pop()
#
#     def top(self):
#         return self.stack[-1]
#
#     def size(self):
#         return len(self.stack)
#
#     def isEmpty(self):
#         return self.stack == []
#
# def countlines(path,count,emitcount):
#     # count = 0
#     # emitcount = 0
#     stack = Stack()
#
#     catalog = open(path)
#     lines = catalog.readlines()
#     i = 0
#     while i < len(lines):
#         # line = lines[i].strip()
#         lines[i] = lines[i].strip()
#         # 如果是空行直接跳过
#         if lines[i] == "":
#             print(i)
#             i += 1
#             continue
#         #遇到emit则emit计数器与总计数器都加一，同时直接略过下面的判断语句
#         if lines[i].startswith("emit "):
#             emitcount[0] += 1
#             count[0] += 1
#             i += 1
#
#         # 遇到注释，如果行首是//那么直接跳过这行，如果行首是/*需要找到*/，在这期间的内容都不要
#         if lines[i].startswith("//"):
#             i += 1
#             continue
#         if lines[i].startswith("/*"):
#             rightzhushi = lines[i].find("*/")
#             # 找到另一半注释
#             while rightzhushi == -1:
#                 i += 1
#                 rightzhushi = lines[i].find("*/")
#             i += 1
#             continue
#
#         # 进行括号处理
#         leftb = lines[i].find("{")
#         rightb = lines[i].find("}")
#         leftm = lines[i].find("[")
#         rightm = lines[i].find("]")
#         lefts = lines[i].find("(")
#         rights = lines[i].find(")")
#         # 如果在一行中没有括号，直接总计数器加一并进行下一行的判断
#         if leftm == lefts == leftb == rights == rightm == rightb == -1:
#             count[0] += 1
#             i += 1
#             continue
#         # 假设程序能通过编译，即每一个括号是匹配的，那么每当遇到左括号将统计代码行数的计数器加一，遇到右括号则直接忽略
#         # 一行中只要有左大括号总计数器就加一
#         if leftb != -1:
#             count[0] += 1
#             i += 1
#         # 在一行中只有右大括号没有左括号则直接跳过，总计数器也不增加
#         elif lines[i].startswith("}"):
#             i += 1
#             continue
#         # 一行中除了右括号还有别的内容
#         else :
#             i += 1
#             count[0] += 1
#         # 所有中括号里的内容都看作一行
#         # 一行中只要有左中括号总计数器就加一
#         if leftm != -1:
#             count[0] += 1
#             # 找到右中括号，在此期间的所有内容都算做一行
#             # 找到右括号  左右括号不在同一行中
#             while rightm == -1:
#                 rightm = lines[i].find("]")
#                 i += 1
#             i += 1
#
#         # 所有小括号中的内容都看作一行
#         # 一行中只要有左小括号总计数器就加一
#         if lefts != -1:
#             count[0] += 1
#             # 找到右小括号，在此期间的所有内容都算做一行
#             # 找到右括号,左右括号不在同一行中
#             while rights == -1:
#                 rights = lines[i].find(")")
#                 i += 1
#             i += 1

def countlines(path,count,emitcount):
    # count = 0
    # emitcount = 0
    # stack = Stack()

    catalog = open(path,"r",encoding="utf-8",errors="ignore")
    lines = catalog.readlines()
    i = 0
    while i < len(lines):
        # line = lines[i].strip()
        lines[i] = lines[i].strip()
        # 如果是空行直接跳过
        if lines[i] == "":
            i += 1
            continue
        #遇到emit则emit计数器与总计数器都加一，同时直接略过下面的判断语句
        if lines[i].startswith("emit "):
            emitcount[0] += 1
            count[0] += 1
            i += 1
            continue

        # 遇到注释，如果行首是//那么直接跳过这行，如果行首是/*需要找到*/，在这期间的内容都不要
        if lines[i].startswith("//"):
            i += 1
            continue
        if lines[i].startswith("/*"):
            rightzhushi = lines[i].find("*/")
            # 找到另一半注释
            while rightzhushi == -1:
                i += 1
                if i >= len(lines):#防止有的人只写注释的前一半，不写后一半
                    break
                rightzhushi = lines[i].find("*/")
            i += 1
            continue

        if lines[i].find(";") != -1 or lines[i].find("{") != -1:
            count[0] += 1
            i += 1
        else :
            i += 1


# 每个项目需要单独调用该函数
def traversal(path,count,emitcount):
    # 首先遍历当前目录所有文件及文件夹
    file_list = os.listdir(path)
    # 循环判断每个元素是否是文件夹还是文件，是文件夹的话，递归
    for file in file_list:
        # 利用os.path.join()方法取得路径全名，并存入cur_path变量，否则每次只能遍历一层目录
        cur_path = os.path.join(path, file)
        # 判断是否是文件夹
        if os.path.isdir(cur_path):
            if os.path.islink(file) or file == 'to_outside':#判断这个文件夹是不是一个软连接，软链接可能导致死循环，只有一个项目中有软链接，名称为to_outside所以直接进行判断了
                continue
            traversal(cur_path,count,emitcount)
        # 判断是否是solidity文件
        elif cur_path.find(".sol",len(cur_path)-4,len(cur_path)) != -1:
            print("1" + cur_path)
            countlines(cur_path,count,emitcount)
            print(count[0],emitcount[0])
        #既不是文件夹也不是solidity文件
        else:
            continue

def countstar(path):
    file_list = os.listdir(path)
    wb = load_workbook("/home/liu/PycharmProjects/pythonProject2/amount.xlsx")
    ws = wb.active

    for file in file_list:
        cur_path = os.path.join(path,file)
        content = open(cur_path,"r")
        strings = content.read()
        jsonStr = json.loads(strings)
        jsonStr['stargazers_count'] #收藏数
        namebefore = jsonStr['full_name']#用户名 + “/” + 仓库名 需要将两个名之间的“/”改为空格才能再表格中找到对因的文件
        nameafter = namebefore.replace('/',' ')
        print(file + " " + nameafter)
        for i in range(2,2936):
            if nameafter == ws.cell(i,1).value:# excel表格的第一行第一列是0，0
                ws.cell(i,6).value = jsonStr['stargazers_count']
                print(i)
                continue

    wb.save("/home/liu/PycharmProjects/pythonProject2/amount.xlsx")


def changepercent(path):
    all_dir = os.listdir(path)
    wb = op.Workbook()
    ws = wb['Sheet']
    ws.append(['项目名','总改动次数','emit改动次数','emit改动次数/总改动次数'])
    for dir in all_dir:
        count = 0#统计一共有多少次改动
        changecount = 0#统计一共有多少次emit改动
        tmp_path = os.path.join(path,dir)
        if os.path.isdir(tmp_path):#进入项目
            cur_path = os.path.join(path,dir)
            file_list = os.listdir(cur_path)
            for file in file_list:
                count += 1#每有一个文件就有一次改动
                print(1)
                if(isEmitChange(tmp_path+"/"+file)):#有emit的改动
                    changecount += 1

        d = dir,count,changecount,((changecount/count) if (count!=0) else -1)
        ws.append(d)

    wb.save("D:/Tool/Tool/PyCharm Community Edition 2022.2/Code/case/emitchange.xlsx")


def isEmitChange(path):
    content = open(path, "r")
    strings = content.read()
    jsonStr = json.loads(strings)
    #如果文件中没有包含一下内容说明emit不会有修改
    if ('files' not in jsonStr):
        return False
    for i in range(len(jsonStr['files'])):
        if 'patch' not in jsonStr['files'][i]:
            continue
        tmpStr = jsonStr['files'][i]['patch'].replace(' ', '')
        lines = tmpStr.split('\n')#将文件修改的patch信息用列表lines来存储

        #开始遍历每一行
        i = 0
        while i < len(lines):
            # 遇到注释，如果行首是//那么直接跳过这行，如果行首是/*需要找到*/，在这期间的内容都不要
            if lines[i].startswith("//") or lines[i].startswith("+//") or lines[i].startswith("-//"):
                i += 1
                continue
            if lines[i].startswith("/*") or lines[i].startswith("+/*") or lines[i].startswith("-/*"):
                rightzhushi = lines[i].find("*/")
                # 找到另一半注释
                while rightzhushi == -1:
                    i += 1
                    if i >= len(lines):  # 防止有的人只写注释的前一半，不写后一半
                        break
                    rightzhushi = lines[i].find("*/")
                i += 1
                continue
            #如果修改的地方有emit，则返回True
            if lines[i].startswith("+emit") or lines[i].startswith("-emit"):
                return True
            #什么都没有发生 到下一行
            i += 1
    # 遍历完所有内容都没有找到修改emit
    return False











